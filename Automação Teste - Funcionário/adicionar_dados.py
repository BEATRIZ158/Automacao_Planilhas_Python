from openpyxl import Workbook, load_workbook

def criar_planilha():
    # Cria uma nova planilha
    medico_enfermeiro = Workbook()

    # Cria uma nova folha (sheet)
    folha = medico_enfermeiro.active

    folha.append(["Nome", "Idade", "Cargo", "Horas Trabalhadas", "Valor da Hora"])

    medico_enfermeiro.save("Planilha_calculo.xlsx")
    
def encontrar_ultima_linha_preenchida(sheet):
    for row in range(1, sheet.max_row + 1):
        if all(not sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            return row
    return sheet.max_row + 1

def adicionar_dados(nome, idade, cargo, qtdHoras, valorHora):
    caminho_planilha = "C:/Users/beatr/OneDrive/Documentos/Projetos Python/Planilha Funcionário/Planilha_calculo.xlsx"

    try:
        medico_enfermeiro = load_workbook(filename=caminho_planilha)
    except FileNotFoundError:
        criar_planilha()
        medico_enfermeiro = load_workbook(filename=caminho_planilha)

    folha = medico_enfermeiro.active
    ultima_linha = encontrar_ultima_linha_preenchida(folha)
    folha.append([nome, idade, cargo, qtdHoras, valorHora])
    medico_enfermeiro.save(caminho_planilha)
