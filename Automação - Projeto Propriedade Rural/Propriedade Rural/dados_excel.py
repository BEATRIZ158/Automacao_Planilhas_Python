import openpyxl

def salvar_em_excel(animal):
    try:
        wb = openpyxl.load_workbook('dados_animais.xlsx')
        ws = wb.active
        
        # Verificar se as colunas já foram adicionadas
        if ws.max_row == 1 and ws.max_column == 1 and ws['A1'].value is None:
            ws.append(["Número do Brinco", "Peso", "Data de Entrada", "Número do Piquet", "Preço Ração (Kg)", "Preço Silo (Kg)", "Ração", "Silo", "Nome do Medicamento", "Valor do Medicamento"])
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Número do Brinco", "Peso", "Data de Entrada", "Número do Piquet", "Preço Ração (Kg)", "Preço Silo (Kg)", "Ração", "Silo", "Nome do Medicamento", "Valor do Medicamento"])
    
    # Adicionar os dados do animal
    ws.append([animal.numero_brinco, animal.peso, animal.data_entrada, animal.numero_piquet, animal.preco_racao, animal.preco_silo, animal.racao, animal.silo, animal.nome_medicamento, animal.valor_medicamento])
    
    wb.save('Automação - Projeto Propriedade Rural/Propriedade Rural/dados_animais.xlsx')
